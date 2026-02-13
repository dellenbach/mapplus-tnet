#!/usr/bin/env python3
"""
excel_to_bookmark.py
Konvertiert Table2.xlsx in bookmark.json-Dateien

Spalten in Excel:
- Webkarte: Bookmarkname oder Aliases
- Pfad_ArcGIS_Server: Layers (komma-separiert)
- Optional: basemap, theme, subtheme

@version    1.0
@date       2026-02-12
@copyright  Trigonet AG
@author     Marco Dellenbach
"""

import openpyxl
import json
import sys
import re
from pathlib import Path

def make_machine_readable(text):
    """
    Normalisiert Texte für Layer-Namen (aus ags2mapplus_config.py)
    - Kleinbuchstaben
    - Umlaute zu ae/oe/ue/ss
    - Nur alphanumerisch, _, /
    - Mehrfach-Unterstriche zu einfach
    """
    text = text.lower()
    text = text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    text = re.sub(r"[^a-z0-9_/]", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")

def read_excel_to_bookmarks(excel_file):
    """
    Liest Excel-Datei und erstellt eine Liste von Bookmarks
    Fasst mehrere Einträge mit gleichem Basis-Namen zusammen
    """
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Header lesen
        headers = {}
        for idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value.lower()] = idx
        
        print(f"Spalten gefunden: {list(headers.keys())}")
        
        # Daten sammeln nach Basis-Namen
        bookmark_groups = {}
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            row_data = {}
            for header, col_idx in headers.items():
                cell = row[col_idx - 1]
                row_data[header] = cell.value
            
            if not row_data.get('webkarte'):
                continue
            
            webkarte = str(row_data.get('webkarte', '')).strip()
            base_name, alias = extract_base_name_and_alias(webkarte)
            
            # Gruppieren nach Basis-Namen
            if base_name not in bookmark_groups:
                bookmark_groups[base_name] = {
                    'base_name': base_name,
                    'aliases': set(),
                    'layers': set(),
                    'basemap': None,
                    'theme': None,
                    'subtheme': None
                }
            
            group = bookmark_groups[base_name]
            
            # Alias hinzufügen
            if alias:
                group['aliases'].add(alias)
            
            # Layers hinzufügen
            layers_raw = row_data.get('pfad_arcgis_server', '')
            if layers_raw:
                layers_str = str(layers_raw).strip()
                if ';' in layers_str:
                    layers = [l.strip() for l in layers_str.split(';') if l.strip()]
                else:
                    layers = [l.strip() for l in layers_str.split(',') if l.strip()]
                
                # URLs verarbeiten wie in ags2mapplus_config.py
                # https://www.geohost.ch/svc/rest/services/gis_fach/nw_verkehrsrichtplan/MapServer/layer1
                # -> servicename: gis_fach/nw_verkehrsrichtplan
                # -> layer: layer1
                # -> layers: gis_fach/nw_verkehrsrichtplan, gis_fach/nw_verkehrsrichtplan/layer1
                processed_layers = set()
                
                for l in layers:
                    # URL-Präfix entfernen
                    url_prefix = 'https://www.geohost.ch/svc/rest/services/'
                    if l.startswith(url_prefix):
                        l = l[len(url_prefix):]
                    
                    # /MapServer/ entfernen (case-insensitive)
                    import re
                    l = re.sub(r'/mapserver/', '/', l, flags=re.IGNORECASE)
                    
                    # Layer-Namen normalisieren (wie ags2mapplus_config.py)
                    full_path = make_machine_readable(l)
                    
                    # Service-Pfad extrahieren (alles vor dem letzten Slash)
                    if '/' in full_path:
                        service_path = full_path.rsplit('/', 1)[0]
                        # Service-Pfad als eigenen Layer hinzufügen
                        processed_layers.add(service_path)
                    
                    # Vollständiger Layer-Pfad hinzufügen
                    processed_layers.add(full_path)
                
                group['layers'].update(processed_layers)
            
            # Andere Felder (erste nicht-leere Werte übernehmen)
            if not group['basemap'] and row_data.get('basemap'):
                group['basemap'] = str(row_data.get('basemap')).strip()
            if not group['theme'] and row_data.get('theme'):
                group['theme'] = str(row_data.get('theme')).strip()
            if not group['subtheme'] and row_data.get('subtheme'):
                group['subtheme'] = str(row_data.get('subtheme')).strip()
        
        # Bookmarks aus Gruppen erstellen
        bookmarks = []
        for base_name, group in sorted(bookmark_groups.items()):
            bookmark = {
                'map-bookmark': group['base_name'],
            }
            
            if group['aliases']:
                bookmark['aliases'] = sorted(list(group['aliases']))
            
            # Basemap setzen (default oder aus Daten)
            bookmark['basemap'] = group['basemap'] if group['basemap'] else 'av_sw'
            
            if group['layers']:
                bookmark['layers'] = sorted(list(group['layers']))
            
            if group['theme']:
                bookmark['theme'] = group['theme']
            
            if group['subtheme']:
                bookmark['subtheme'] = group['subtheme']
            
            bookmarks.append(bookmark)
            alias_str = f" + Aliases: {', '.join(sorted(group['aliases']))}" if group['aliases'] else ""
            print(f"✓ Bookmark erstellt: {bookmark['map-bookmark']}{alias_str}")
        
        return bookmarks
    
    except FileNotFoundError:
        print(f"Fehler: Datei {excel_file} nicht gefunden!")
        sys.exit(1)
    except Exception as e:
        print(f"Fehler beim Lesen der Excel-Datei: {e}")
        sys.exit(1)

def extract_base_name_and_alias(webkarte):
    """
    Extrahiert den Basis-Namen und den Alias aus einem Webkarte-Eintrag
    z.B. "map_name_intern" -> ("map_name", "map_name_intern")
    Normalisiert beide Namen mit make_machine_readable
    """
    name = webkarte.strip()
    
    # Bekannte Suffixe für Aliases
    suffixes = ['_intern', '_public', '_export', '_sandbox', '_dev', '_test']
    
    for suffix in suffixes:
        if name.endswith(suffix):
            base = name[:-len(suffix)]
            return (make_machine_readable(base), make_machine_readable(name))
    
    # Kein bekannter Suffix -> Name ist selbst der Basis-Name
    return (make_machine_readable(name), None)

def create_bookmark(row_data):
    """
    Diese Funktion wird nicht mehr verwendet - Logik wurde in read_excel_to_bookmarks integriert
    """
    pass

def save_bookmarks(bookmarks, output_file):
    """
    Speichert Bookmarks als JSON-Datei
    """
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(bookmarks, f, indent=2, ensure_ascii=False)
        print(f"\n✓ Bookmarks gespeichert in: {output_file}")
        print(f"  Anzahl Bookmarks: {len(bookmarks)}")
    except Exception as e:
        print(f"Fehler beim Speichern der JSON-Datei: {e}")
        sys.exit(1)

def main():
    # Pfade
    script_dir = Path(__file__).parent
    excel_file = script_dir / 'Table2.xlsx'
    output_file = script_dir / 'map-bookmarks-all.json'
    
    print("=" * 60)
    print("Excel zu Bookmark JSON Konverter")
    print("=" * 60)
    print(f"\nEingabe:  {excel_file}")
    print(f"Ausgabe:  {output_file}\n")
    
    # Excel lesen
    bookmarks = read_excel_to_bookmarks(excel_file)
    
    if not bookmarks:
        print("\n⚠ Keine Bookmarks gefunden!")
        sys.exit(1)
    
    # JSON speichern
    save_bookmarks(bookmarks, output_file)
    
    print("\n✓ Konvertierung erfolgreich!")

if __name__ == '__main__':
    main()
